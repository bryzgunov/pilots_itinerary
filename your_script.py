import fitz
import unicodedata
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from PIL import Image as PILImage
import io
import os
import numpy as np
import re

# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===
def normalize_ascii(text):
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if ord(c) < 128)

def is_takeoff_file(pdf_bytes):
    """Определяет, содержит ли файл 'Takeoff' в начале"""
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        raw = doc[0].get_text("text")[:250]
        doc.close()
        return normalize_ascii(raw).strip().lower().startswith("takeoff")
    except Exception as e:
        print(f"Ошибка при проверке файла на Takeoff: {e}")
        return False

def extract_first_n_lines_from_doc(doc, n=32):
    page = doc[0]
    blocks = page.get_text("dict")["blocks"]
    blocks.sort(key=lambda b: (b["bbox"][1], b["bbox"][0]))
    lines = []
    for block in blocks:
        if "lines" not in block:
            continue
        for line in block["lines"]:
            text = "".join(span["text"] for span in line["spans"]).strip()
            if text:
                lines.append(text)
                if len(lines) >= n:
                    return lines
    return lines

def parse_document_with_simple_split(page, target_phrase="All Engines Operating"):
    phrase_y_coord = None
    text_instances = page.search_for(target_phrase)

    if text_instances:
        phrase_y_coord = text_instances[0].y0
    else:
        print(f"⚠️ Фраза '{target_phrase}' не найдена в документе.")
        return [], []

    page_width = page.rect.width
    mid_x = page.rect.x0 + page_width / 2

    blocks = page.get_text("dict")["blocks"]

    left_array = []
    right_array = []

    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                if line["bbox"][1] < phrase_y_coord:
                    line_text = "".join(span["text"] for span in line["spans"]).strip()
                    if line_text:
                        max_x = max(span["bbox"][2] for span in line["spans"])
                        if max_x < mid_x:
                            left_array.append(line_text)
                        else:
                            right_array.append(line_text)
                else:
                    continue

    return left_array, right_array

def extract_variables(arr, suffix):
    variables = {}

    runway_count = 0
    for i, item in enumerate(arr):
        if "Runway" in item:
            runway_count += 1
            if runway_count == 2 and i + 1 < len(arr):
                variables[f"Runway{suffix}"] = arr[i + 1]
                break

    for i, item in enumerate(arr):
        if "Usable Length" in item and i + 1 < len(arr):
            variables[f"Length{suffix}"] = arr[i + 1]
            break

    for i, item in enumerate(arr):
        if "Runway Surface" in item and i + 1 < len(arr):
            variables[f"Surface{suffix}"] = arr[i + 1]
            break

    wind_idx = -1
    temp_idx = -1
    for i, item in enumerate(arr):
        if "Wind" in item:
            wind_idx = i
        elif "Temperature" in item:
            temp_idx = i
            break

    if wind_idx != -1 and temp_idx != -1 and wind_idx < temp_idx:
        wind_content = " ".join(arr[wind_idx + 1 : temp_idx])
        variables[f"Wind{suffix}"] = wind_content

    for i, item in enumerate(arr):
        if "Altimeter" in item and i + 1 < len(arr):
            next_item = arr[i + 1]
            if "/" in next_item:
                variables[f"Altimeter{suffix}"] = next_item.split("/", 1)[1].strip()
            else:
                variables[f"Altimeter{suffix}"] = next_item
            break

    for i, item in enumerate(arr):
        if "Distance" in item and "Safety Distance Factor" not in item and i + 1 < len(arr):
            next_item = arr[i + 1]
            if "/" in next_item:
                variables[f"Distance{suffix}"] = next_item.split("/", 1)[1].strip()
            else:
                variables[f"Distance{suffix}"] = next_item
            break

    return variables

def process_runway_variable(runway_value, suffix):
    if not runway_value:
        return runway_value, runway_value

    runway0 = runway_value.strip()

    if re.match(r'^\d+$', runway0):
        num = int(runway0)
        if num < 18:
            new_runway = f"{num:02d}/{num+18:02d}"
        else:
            new_runway = f"{num-18:02d}/{num:02d}"
        return runway0, new_runway

    elif re.match(r'^\d+[LR]$', runway0):
        num_part = re.findall(r'\d+', runway0)[0]
        letter_part = re.findall(r'[LR]', runway0)[0]
        runway0_numeric = num_part
        opposite_letter = 'L' if letter_part == 'R' else 'R'
        num = int(num_part)
        if num < 18:
            new_runway = f"{num_part}{letter_part}/{num+18:02d}{opposite_letter}"
        else:
            new_runway = f"{num-18:02d}{opposite_letter}/{num_part}{letter_part}"
        return runway0_numeric, new_runway

    elif '/' in runway0:
        before_slash = runway0.split('/')[0]
        numeric_part = re.findall(r'\d+', before_slash)
        if numeric_part:
            runway0_numeric = numeric_part[0]
        else:
            runway0_numeric = before_slash
        new_runway = runway0
        return runway0_numeric, new_runway

    else:
        return runway0, runway0

def process_wind_variable(wind_value, runway0_value, suffix):
    if not wind_value:
        return wind_value, wind_value

    wind0 = wind_value.strip()

    kts_matches = list(re.finditer(r'(\d+(?:-\d+)?)\s*kts', wind0, re.IGNORECASE))
    degree_match = re.search(r'(\d+)°T', wind0)

    if len(kts_matches) >= 3 and degree_match:
        wind_x_1_full = kts_matches[1].group(1)
        wind_x_1 = int(wind_x_1_full.split('-')[-1])

        wind_x_2_full = kts_matches[2].group(1)
        wind_x_2 = int(wind_x_2_full.split('-')[-1])

        wind_x_3 = degree_match.group(1) + "°"

        wind_x_4_full = kts_matches[0].group(1)
        wind_x_4 = int(wind_x_4_full.split('-')[-1])

        try:
            runway0_num = int(runway0_value)
        except ValueError:
            runway0_numeric_match = re.search(r'\d+', str(runway0_value))
            if runway0_numeric_match:
                runway0_num = int(runway0_numeric_match.group())
            else:
                return wind0, wind0

        abs_val_1 = abs(runway0_num * 10 - wind_x_1)
        if 90 <= abs_val_1 <= 270:
            wind_x_1_str = "H" + str(wind_x_1)
        else:
            wind_x_1_str = "T" + str(wind_x_1)

        abs_val_2 = abs(runway0_num * 10 - wind_x_2)
        if abs_val_2 > 0:
            if (runway0_num * 10 - wind_x_2) > 0:
                wind_x_2_str = "L" + str(wind_x_2)
            else:
                wind_x_2_str = "R" + str(wind_x_2)
        else:
            wind_x_2_str = "L" + str(wind_x_2)

        new_wind = f"{wind_x_1_str}/{wind_x_2_str} ({wind_x_3}/{wind_x_4})"

        return wind0, new_wind
    else:
        return wind0, wind0

def parse_main_route_table(doc):
    """Парсит основную таблицу маршрута из PDF"""
    page = doc[0]
    all_words = page.get_text("words")

    # --- Поиск заголовка ---
    target_y = None
    for word_tuple in all_words:
        x0, y0, x1, y1, text, *_ = word_tuple
        if text == "WAYPOINT":
            for w in all_words:
                wx0, wy0, wx1, wy1, wtext, *_ = w
                if wtext == "ACT" and abs((y0 + y1)/2 - (wy0 + wy1)/2) < 5 and wx0 > x0:
                    target_y = (y0 + y1) / 2
                    break
            if target_y is not None:
                break

    if target_y is None:
        for word_tuple in all_words:
            x0, y0, x1, y1, text, *_ = word_tuple
            if text == "MAG":
                target_y = (y0 + y1) / 2 + 15
                break

    if target_y is None:
        raise RuntimeError("Не найдена строка заголовка таблицы.")

    # --- Извлечение заголовков ---
    header_keywords = ["WAYPOINT", "AIRWAY", "HDG", "CRS", "ALT", "CMP", "DIR/SPD", "ISA", "TAS", "GS", "LEG", "REM", "USED", "ACT", "ETE"]
    header_words_info = []
    tolerance = 5.0
    for word_tuple in all_words:
        x0, y0, x1, y1, text, *_ = word_tuple
        center_y = (y0 + y1) / 2
        if abs(center_y - target_y) <= tolerance and text in header_keywords:
            header_words_info.append((text, x0, x1))

    header_words_info.sort(key=lambda item: item[1])

    # --- Построение XX ---
    XX = []
    for i in range(1, len(header_words_info)):
        x1_prev = header_words_info[i-1][2]
        x0_next = header_words_info[i][1]
        boundary_x = (x0_next - x1_prev) / 2 + x1_prev
        XX.append(boundary_x)

    if XX:
        x0_airway = next((x0 for text, x0, x1 in header_words_info if text == "AIRWAY"), None)
        if x0_airway is not None:
            XX[0] = x0_airway - 2
        XX.insert(0, 5)
        XX.append(XX[-1] + 10)

    # --- Построение YY ---
    alt_coords = None
    for text, x0, x1 in header_words_info:
        if text == "ALT":
            for wx0, wy0, wx1, wy1, wtext, *_ in all_words:
                if wtext == "ALT" and abs(wx0 - x0) < 1 and abs(wx1 - x1) < 1:
                    alt_coords = (wx0, wy0, wx1, wy1)
                    break
            if alt_coords:
                break

    if not alt_coords:
        raise RuntimeError("Не найдены координаты слова 'ALT'.")

    x0_alt, y0_alt, x1_alt, y1_alt = alt_coords

    y0_alternate = None
    for wx0, wy0, wx1, wy1, wtext, *_ in all_words:
        if "ALTERNATE" in wtext:
            y0_alternate = wy0
            break
    if y0_alternate is None:
        for wx0, wy0, wx1, wy1, wtext, *_ in all_words:
            if "2000 FT" in wtext and "ISA:" in wtext:
                y0_alternate = wy0
                break

    if y0_alternate is None:
        raise RuntimeError("Не найдена нижняя граница таблицы.")

    YY = []
    for wx0, wy0, wx1, wy1, wtext, *_ in all_words:
        if x0_alt <= (wx0 + wx1) / 2 <= x1_alt and y1_alt <= wy0 <= y0_alternate:
            if wtext != "ALT" and "ALTERNATE" not in wtext and "2000 FT" not in wtext:
                YY.append(wy0 - 2)
    YY.append(y0_alternate - 2)

    # --- Парсинг сетки ---
    num_cols = len(XX) - 1
    num_rows = len(YY) - 1

    exact_columns = [
        "WAYPOINT", "AIRWAY", "HDG", "CRS", "ALT", "CMP", "DIR/SPD", "ISA",
        "TAS", "GS", "LEG", "REM", "USED", "REM", "ACT", "LEG", "REM", "ETE", "ACT"
    ]

    data_grid = []
    if num_cols <= len(exact_columns):
        for row_idx in range(num_rows):
            row_data = [''] * len(exact_columns)
            for col_idx in range(min(num_cols, len(exact_columns))):
                x_min = XX[col_idx]
                x_max = XX[col_idx + 1]
                y_min = YY[row_idx]
                y_max = YY[row_idx + 1]
                cell_texts = []
                for word_tuple in all_words:
                    wx0, wy0, wx1, wy1, wtext, *_ = word_tuple
                    center_x = (wx0 + wx1) / 2
                    center_y = (wy0 + wy1) / 2
                    if x_min <= center_x <= x_max and y_min <= center_y <= y_max:
                        cell_texts.append(wtext)
                row_data[col_idx] = ' '.join(cell_texts) if cell_texts else ''
            data_grid.append(row_data)
    else:
        for row_idx in range(num_rows):
            row_data = []
            for col_idx in range(len(exact_columns)):
                x_min = XX[col_idx]
                x_max = XX[col_idx + 1]
                y_min = YY[row_idx]
                y_max = YY[row_idx + 1]
                cell_texts = []
                for word_tuple in all_words:
                    wx0, wy0, wx1, wy1, wtext, *_ = word_tuple
                    center_x = (wx0 + wx1) / 2
                    center_y = (wy0 + wy1) / 2
                    if x_min <= center_x <= x_max and y_min <= center_y <= y_max:
                        cell_texts.append(wtext)
                row_data.append(' '.join(cell_texts) if cell_texts else '')
            data_grid.append(row_data)

    df = pd.DataFrame(data_grid, columns=exact_columns)
    return df

def parse_airport_table(doc):
    """Парсит таблицу аэропортов"""
    target_text = "AIRPORT"
    page_with_table = None
    
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        text = page.get_text()
        if target_text in text:
            page_with_table = page
            print(f"Ключевое слово '{target_text}' найдено на странице {page_num + 1}")
            break

    if page_with_table is None:
        print(f"Ключевое слово '{target_text}' не найдено.")
        return pd.DataFrame()

    # Поиск координат "AIRPORT"
    words = page_with_table.get_text("words")
    airport_coords = None
    for word in words:
        if word[4].upper() == target_text:
            airport_coords = (word[0], word[1], word[2], word[3])
            break

    if airport_coords is None:
        print("Не удалось получить координаты слова 'AIRPORT'.")
        return pd.DataFrame()

    # Определяем XX
    XX_airport = [5, 75, 150, 200, 250, 325, 375, 425, 475, 525, 600]

    # Определяем YY
    YY_airport = []
    YY_airport.append(airport_coords[3] + 2)

    dest_coords = None
    for word in words:
        if word[4].upper() == "DEST" and word[1] > airport_coords[3]:
            dest_coords = (word[0], word[1], word[2], word[3])
            break

    if dest_coords is None:
        # Используем приблизительные координаты
        words_below_airport = [w for w in words if w[1] > airport_coords[3]]
        if words_below_airport:
            min_y0_below = min([w[1] for w in words_below_airport])
            avg_height = 15
            YY_airport.append(min_y0_below - 2)
            YY_airport.append(min_y0_below + avg_height + 2)
        else:
            YY_airport.append(airport_coords[3] + 30)
            YY_airport.append(airport_coords[3] + 50)
    else:
        YY_airport.append(dest_coords[1] - 2)
        YY_airport.append(dest_coords[3] + 2)

    # Извлечение текста
    blocks = page_with_table.get_text("dict").get("blocks", [])
    extracted_text_dict = {}
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text_content = span["text"].strip()
                    bbox = span["bbox"]
                    center_x = (bbox[0] + bbox[2]) / 2
                    center_y = (bbox[1] + bbox[3]) / 2
                    extracted_text_dict[(center_x, center_y)] = text_content

    def find_text_in_rect(x0, y0, x1, y1):
        found_texts = []
        for (cx, cy), text in extracted_text_dict.items():
            if x0 <= cx <= x1 and y0 <= cy <= y1:
                found_texts.append(text)
        return " ".join(found_texts).strip()

    num_cols_airport = len(XX_airport) - 1
    num_rows_airport = len(YY_airport) - 1
    df_data_airport = []

    for j in range(num_rows_airport):
        row_data = []
        for i in range(num_cols_airport):
            x_start = XX_airport[i]
            x_end = XX_airport[i+1]
            y_start = YY_airport[j]
            y_end = YY_airport[j+1]
            cell_text = find_text_in_rect(x_start, y_start, x_end, y_end)
            row_data.append(cell_text)
        df_data_airport.append(row_data)

    return pd.DataFrame(df_data_airport)

def extract_airport_maps(doc):
    """Извлекает карты аэропортов с последней страницы"""
    last_page = doc[-1]
    
    # Извлечение текста построчно
    text_blocks = last_page.get_text("dict")
    lines = []
    for block in text_blocks["blocks"]:
        if "lines" in block:
            for line in block["lines"]:
                line_text = "".join(span["text"] for span in line["spans"])
                stripped = line_text.strip()
                if stripped:
                    lines.append(stripped)

    if len(lines) < 4:
        text_A1 = "DEP"
        text_A28 = "DEST"
    else:
        text_A1 = lines[1] if len(lines) > 1 else "DEP"
        text_A28 = lines[3] if len(lines) > 3 else "DEST"

    # Извлечение и обработка изображений
    image_list = last_page.get_images(full=True)
    img_buffers = []

    for idx, img in enumerate(image_list):
        try:
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            pil_img = PILImage.open(io.BytesIO(image_bytes))
            pil_img = pil_img.resize((500, 500), PILImage.LANCZOS)
            
            # Сохраняем в буфер памяти
            img_buffer = io.BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            img_buffers.append(img_buffer)
        except Exception as e:
            print(f"Ошибка при обработке изображения {idx}: {e}")

    return text_A1, text_A28, img_buffers

def extract_screenshot_for_generated_sheet(doc):
    """Извлекает скриншот для вставки в Generated_Sheet"""
    page = doc[0]
    
    # Получаем все спаны на странице
    blocks = page.get_text("dict")["blocks"]
    spans = []
    for block in blocks:
        if "lines" in block:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"].strip()
                    if text:
                        bbox = span["bbox"]
                        spans.append({
                            "text": text,
                            "x0": bbox[0],
                            "y0": bbox[1],
                            "x1": bbox[2],
                            "y1": bbox[3]
                        })

    if not spans:
        print("⚠️ Не найдено текстовых элементов на странице")
        return None

    # Сортируем спаны
    spans.sort(key=lambda s: (s["y0"], s["x0"]))

    # 1. Самое первое слово на листе
    first_span = spans[0]
    x01 = first_span["x0"]
    y01 = first_span["y0"]

    # 2. Поиск первого вхождения "Route"
    y02 = None
    for span in spans:
        if re.search(r'\bRoute\b', span["text"], re.IGNORECASE):
            y02 = span["y0"]
            break

    # 3. Поиск первого вхождения "Landing Fuel"
    x02 = None
    for span in spans:
        if re.search(r'\bLanding\s+Fuel\b', span["text"], re.IGNORECASE):
            x02 = span["x1"]
            break
        elif re.search(r'\bLanding\b', span["text"], re.IGNORECASE):
            idx = spans.index(span)
            for next_span in spans[idx+1:min(idx+5, len(spans))]:
                if abs(next_span["y0"] - span["y0"]) < 5 and re.search(r'\bFuel\b', next_span["text"], re.IGNORECASE):
                    x02 = next_span["x1"]
                    break
            if x02:
                break

    if y02 is None or x02 is None:
        print("⚠️ Не все необходимые координаты найдены")
        return None

    # Вычисляем координаты области для вырезания
    clip_rect = fitz.Rect(
        x01 - 5,
        y01 - 3,
        x02 + 30,
        y02 - 15
    )

    # Проверка корректности прямоугольника
    if clip_rect.x0 < 0: clip_rect.x0 = 0
    if clip_rect.y0 < 0: clip_rect.y0 = 0
    if clip_rect.x1 > page.rect.width: clip_rect.x1 = page.rect.width
    if clip_rect.y1 > page.rect.height: clip_rect.y1 = page.rect.height

    if clip_rect.is_empty or clip_rect.get_area() < 1:
        print(f"⚠️ Некорректная область вырезания: {clip_rect}")
        return None

    # Рендерим страницу и вырезаем область
    pix = page.get_pixmap(dpi=150, clip=clip_rect)
    img_data = pix.tobytes("png")
    
    # Создаем PIL Image
    image = PILImage.open(io.BytesIO(img_data))
    
    # Сохраняем в буфер
    img_buffer = io.BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    
    return img_buffer

def create_generated_sheet(wb, ws1, ws2, ws3, ws4, ws5, screenshot_buffer=None):
    """Создает лист Generated_Sheet"""
    new_sheet_name = "Generated_Sheet"
    if new_sheet_name in wb.sheetnames:
        wb.remove(wb[new_sheet_name])
    ws = wb.create_sheet(title=new_sheet_name)

    # Общий шрифт
    default_font = Font(name='Helvetica Neue', size=11)

    # Ширина столбцов
    col_widths = {'A': 5, 'B': 22, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 31, 'I': 35}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Стили
    bold_gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(name='Helvetica Neue', size=11, bold=True)
    header_font = Font(name='Helvetica Neue', size=11, bold=True)

    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ПЕРЕМЕННАЯ СМЕЩЕНИЯ
    offset_rows = 7

    # Убираем границы у первых 7 строк
    for row in range(1, offset_rows + 1):
        for col in range(1, 10):  # Теперь до столбца I
            cell = ws.cell(row=row, column=col)
            cell.border = Border()

    # Вставляем текст в ячейку A7 (offset_rows)
    info_row = offset_rows
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=8)
    info_cell = ws.cell(row=info_row, column=1)
    info_cell.value = "Tacho start: ______ Off Block: ______ Take Off: ______ Tacho end: ______ Landing: ______ On Block: ______"
    info_cell.font = Font(name='Helvetica Neue', size=9)
    info_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Заголовки (строки 8–9)
    header_row_1 = 8
    header_row_2 = 9

    headers = {
        f'A{header_row_1}': ('№', 'center', 'top', True, 2),
        f'B{header_row_1}': ('Waypoint', 'left', 'top', True, 2),
        f'C{header_row_1}': ('ALT', 'center', 'center', True, 2),
        f'D{header_row_1}': ('HDG', 'left', 'center', False, 1),
        f'E{header_row_1}': ('Dist.', 'left', 'center', False, 1),
        f'F{header_row_1}': ('EFOB', 'left', 'center', False, 1),
        f'G{header_row_1}': ('ETA', 'left', 'center', False, 1),
        f'H{header_row_1}': ('Radio', 'left', 'top', True, 2),
        f'D{header_row_2}': ('CRS', 'right', 'center', False, 1),
        f'E{header_row_2}': ('Time', 'right', 'center', False, 1),
        f'F{header_row_2}': ('AFOB', 'right', 'center', False, 1),
        f'G{header_row_2}': ('ATA', 'right', 'center', False, 1),
    }

    for cell_ref, (value, h_align, v_align, merge, rows) in headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=False)
        if merge:
            col_letter = cell_ref[0]
            start_row = int(cell_ref[1:])
            end_row = start_row + rows - 1
            col_idx = ord(col_letter.upper()) - ord('A') + 1
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

    # Фон и жирный шрифт для A8:H9
    for row in range(header_row_1, header_row_1 + 2):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = bold_gray_fill
            cell.font = bold_font
            cell.border = thin_border

    # Переменные
    y0 = 5 + offset_rows
    last_row_main = ws2.max_row
    x = last_row_main - 3

    # Обработка строк
    for i in range(1, x + 1):
        row_offset = y0 + i * 3

        # A: номер с ведущим нулём
        num_val = i + 1
        a_val = f"{num_val:02d}"
        a_cell = ws.cell(row=row_offset, column=1, value=a_val)
        a_cell.font = default_font
        a_cell.alignment = Alignment(horizontal='center', vertical='top')
        if i == x:
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset+4, end_column=1)
        else:
            ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset+2, end_column=1)

        # B: Waypoint
        b_val = ws2.cell(row=i+3, column=1).value
        b_cell = ws.cell(row=row_offset, column=2, value=b_val)
        b_cell.font = default_font
        b_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        if i == x:
            ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset+4, end_column=2)
        else:
            ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset+2, end_column=2)

        # C: ALT
        c_val = ws2.cell(row=i+3, column=5).value
        if c_val:
            c_cell = ws.cell(row=row_offset - 1, column=3, value=c_val)
            c_cell.font = default_font
            c_cell.alignment = Alignment(horizontal='center', vertical='center')
            if i < x:
                ws.merge_cells(start_row=row_offset - 1, start_column=3, end_row=row_offset, end_column=3)

        # D: HDG/CRS
        d1_val = ws2.cell(row=i+3, column=3).value
        if d1_val:
            d1_cell = ws.cell(row=row_offset - 1, column=4, value=d1_val)
            d1_cell.font = default_font
            d1_cell.alignment = Alignment(horizontal='left', vertical='center')

        d2_val = ws2.cell(row=i+3, column=4).value
        if d2_val:
            d2_cell = ws.cell(row=row_offset, column=4, value=d2_val)
            d2_cell.font = default_font
            d2_cell.alignment = Alignment(horizontal='right', vertical='center')

        # E: Dist/Time
        e1_val = ws2.cell(row=i+3, column=11).value
        if e1_val:
            e1_cell = ws.cell(row=row_offset - 1, column=5, value=e1_val)
            e1_cell.font = default_font
            e1_cell.alignment = Alignment(horizontal='left', vertical='center')

        e2_val = ws2.cell(row=i+3, column=16).value
        if e2_val:
            e2_cell = ws.cell(row=row_offset, column=5, value=e2_val)
            e2_cell.font = default_font
            e2_cell.alignment = Alignment(horizontal='right', vertical='center')

        # F: EFOB/AFOB
        f_val = ws2.cell(row=i+3, column=14).value
        if f_val:
            f_cell = ws.cell(row=row_offset - 1, column=6, value=f_val)
            f_cell.font = default_font
            f_cell.alignment = Alignment(horizontal='left', vertical='center')

        # H: Radio
        h_start = row_offset - 1
        h_end = row_offset
        ws.merge_cells(start_row=h_start, start_column=8, end_row=h_end, end_column=8)
        h_cell = ws.cell(row=h_start, column=8)
        h_cell.font = default_font
        h_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Последняя строка блока
        if i < x:
            merge_range = f"C{row_offset + 1}:H{row_offset + 1}"
            ws.merge_cells(merge_range)
            merged_cell = ws.cell(row=row_offset + 1, column=3)
            merged_cell.font = default_font
            merged_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Первый блок
    start_a = 3 + offset_rows
    start_b = 3 + offset_rows

    ws.merge_cells(start_row=start_a, start_column=1, end_row=start_a+4, end_column=1)
    a3 = ws.cell(row=start_a, column=1)
    a3.value = "01"
    a3.font = default_font
    a3.alignment = Alignment(horizontal='center', vertical='top')

    b3_val = ws2.cell(row=3, column=1).value
    ws.merge_cells(start_row=start_b, start_column=2, end_row=start_b+4, end_column=2)
    b3 = ws.cell(row=start_b, column=2)
    b3.value = b3_val
    b3.font = default_font
    b3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.merge_cells(start_row=start_a, start_column=3, end_row=start_a+3, end_column=8)
    c3 = ws.cell(row=start_a, column=3)

    # Формирование текста для C3
    h2_val_raw = ws3['H2'].value if ws3['H2'].value else "0"
    try:
        h2_val = round(float(h2_val_raw))
    except (ValueError, TypeError):
        h2_val = 0

    h2_plus_300 = ((h2_val + 300) // 10 + (1 if (h2_val + 300) % 10 != 0 else 0)) * 10

    # Используем значения с листов
    atis_val = ws3['D2'].value if ws3['D2'].value else "_____"
    gnd_val = ws3['G2'].value if ws3['G2'].value else "_____"
    twr_val = ws3['E2'].value if ws3['E2'].value else "_____"
    
    # Безопасное получение значений из ws5
    exp_rwy = ws5.cell(row=2, column=4).value if ws5.cell(row=2, column=4).value else "_____"
    rwy = ws5.cell(row=3, column=4).value if ws5.cell(row=3, column=4).value else "_____"
    length = ws5.cell(row=4, column=4).value if ws5.cell(row=4, column=4).value else "_____"
    req_dist = ws5.cell(row=9, column=4).value if ws5.cell(row=9, column=4).value else "_____"
    surface = ws5.cell(row=5, column=4).value if ws5.cell(row=5, column=4).value else "_____"
    exp_wind = ws5.cell(row=7, column=4).value if ws5.cell(row=7, column=4).value else "_____"
    exp_qnh = ws5.cell(row=8, column=4).value if ws5.cell(row=8, column=4).value else "_____"

    text_c3 = (
        f"Departure ({ws4['A1'].value if ws4['A1'].value else '____'}, ______,{h2_val}, {h2_plus_300}, _____ , Exp. RWY: {exp_rwy}\n"
        f"ATIS: {atis_val}; GND: {gnd_val}; TWR: {twr_val};\n"
        f"RWY: {rwy}; Length: {length}; Req. Dist.: {req_dist}; Surface: {surface};\n"
        f"Exp. Wind: {exp_wind}; Exp. QNH: {exp_qnh}; Exp. TWY:_____\n"
        f"RWY: ____ ; Wind: ________; QNH: _______; Squak: ________"
    )

    c3.value = text_c3
    c3.font = Font(name='Helvetica Neue', size=9)
    c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Последний блок
    final_start_row = y0 + x * 3 + 1
    final_end_row = final_start_row + 3
    ws.merge_cells(start_row=final_start_row, start_column=3, end_row=final_end_row, end_column=8)
    final_cell = ws.cell(row=final_start_row, column=3)

    # Формирование финального текста
    h3_val_raw = ws3['H3'].value if ws3['H3'].value else "0"
    try:
        h3_val = round(float(h3_val_raw))
    except (ValueError, TypeError):
        h3_val = 0

    h3_plus_300 = ((h3_val + 300) // 10 + (1 if (h3_val + 300) % 10 != 0 else 0)) * 10

    # Безопасное получение значений из ws5
    atis_val_f = ws3['D3'].value if ws3['D3'].value else "_____"
    gnd_val_f = ws3['G3'].value if ws3['G3'].value else "_____"
    twr_val_f = ws3['E3'].value if ws3['E3'].value else "_____"
    
    exp_rwy_f = ws5.cell(row=2, column=6).value if ws5.cell(row=2, column=6).value else "_____"
    rwy_f = ws5.cell(row=3, column=6).value if ws5.cell(row=3, column=6).value else "_____"
    length_f = ws5.cell(row=4, column=6).value if ws5.cell(row=4, column=6).value else "_____"
    req_dist_f = ws5.cell(row=9, column=6).value if ws5.cell(row=9, column=6).value else "_____"
    surface_f = ws5.cell(row=5, column=6).value if ws5.cell(row=5, column=6).value else "_____"
    exp_wind_f = ws5.cell(row=7, column=6).value if ws5.cell(row=7, column=6).value else "_____"
    exp_qnh_f = ws5.cell(row=8, column=6).value if ws5.cell(row=8, column=6).value else "_____"

    text_final = (
        f"Destination ({ws4['A28'].value if ws4['A28'].value else '____'}, ______,{h3_val}, {h3_plus_300}, _____ , Exp. RWY: {exp_rwy_f}\n"
        f"ATIS: {atis_val_f}; GND: {gnd_val_f}; TWR: {twr_val_f};\n"
        f"RWY: {rwy_f}; Length: {length_f}; Req. Dist.: {req_dist_f}; Surface: {surface_f};\n"
        f"Exp. Wind: {exp_wind_f}; Exp. QNH: {exp_qnh_f}; Exp. TWY:_____\n"
        f"RWY: ____ ; Wind: ________; QNH: _______; Squak: ________"
    )

    final_cell.value = text_final
    final_cell.font = Font(name='Helvetica Neue', size=9)
    final_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Настройка высоты строк
    last_output_row = y0 + x * 3 + 4
    for row_num in range(1, last_output_row + 1):
        if (start_a <= row_num <= start_a+3) or (final_start_row <= row_num <= final_end_row):
            ws.row_dimensions[row_num].height = 15
        else:
            ws.row_dimensions[row_num].height = 14

    # Применяем границы и шрифт
    for row in ws.iter_rows(min_row=1, max_row=last_output_row, min_col=1, max_col=8):
        for cell in row:
            if cell.font.size != 9:
                cell.font = default_font
            if cell.row <= offset_rows:
                continue
            cell.border = thin_border

    # Вставляем дополнительный текст
    final_info_row = final_end_row + 2
    info_text = (
        "TEM (Threats error management), CANWE (Crew, Aircraft, Notam, Weather, Environment) \n"
        "After T/O: Flaps, Lights, Engine        Approach: QNH, Mixture, Fuel, Flaps \n"
        "Landing: Mixture, Flaps, Lights         After Landing: Heat, Light, Flaps\n"
        "Waypoint: Top, Track, Altitude, Radio, Engine, Estimates, Area\n"
        "Diversion: Aircraft Endurance, Terrain, Infrastructure, Weather, Airport\n"
        "Arrival Briefing (Treats, RWY, Top Of Descent, Integration, Missed Aproach, Holding time, Landing configuration and speed, Taxiway, Apron)"
    )

    ws.cell(row=final_info_row, column=1, value=info_text)
    ws.cell(row=final_info_row, column=1).font = Font(name='Helvetica Neue', size=8)
    ws.cell(row=final_info_row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=final_info_row, start_column=1, end_row=final_info_row, end_column=8)
    ws.row_dimensions[final_info_row].height = 70

    # Настройка разметки страницы
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3, header=0.1, footer=0.1)

    # ВСТАВКА СКРИНШОТА В СТОЛБЕЦ I (если есть)
    if screenshot_buffer:
        try:
            # Создаем изображение для вставки
            xl_img = XLImage(screenshot_buffer)
            
            # Уменьшаем изображение
            scale_factor = 1.8
            xl_img.width = int(xl_img.width / scale_factor)
            xl_img.height = int(xl_img.height / scale_factor)
            
            # Вставляем изображение в столбец I, строку 1
            # Это НЕ затрёт данные в столбцах A-H
            xl_img.anchor = 'I1'
            ws.add_image(xl_img)
            print("✅ Скриншот вставлен в столбец I листа 'Generated_Sheet'")
        except Exception as e:
            print(f"⚠️ Ошибка при вставке скриншота: {e}")

def process_two_pdfs(file1_bytes, file2_bytes, file1_name, file2_name):
    """
    Основная функция обработки двух PDF файлов
    Вход: два PDF файла в виде байтов
    Выход: Excel файл в виде байтов
    """
    print(f"Начинаю обработку файлов: {file1_name} и {file2_name}")
    
    # Определяем, какой файл является Takeoff
    file1_is_takeoff = is_takeoff_file(file1_bytes)
    file2_is_takeoff = is_takeoff_file(file2_bytes)
    
    print(f"Файл 1 ({file1_name}) содержит Takeoff: {file1_is_takeoff}")
    print(f"Файл 2 ({file2_name}) содержит Takeoff: {file2_is_takeoff}")
    
    # Проверяем, что ровно один файл содержит Takeoff
    if file1_is_takeoff and file2_is_takeoff:
        raise ValueError("Оба файла содержат 'Takeoff'. Нужен только один файл с Takeoff.")
    elif not file1_is_takeoff and not file2_is_takeoff:
        raise ValueError("Ни один из файлов не содержит 'Takeoff'. Нужен один файл с Takeoff.")
    
    # Определяем, какой файл обрабатывать как основной (не Takeoff)
    if file1_is_takeoff:
        main_file_bytes = file2_bytes  # file2 - основной
        takeoff_file_bytes = file1_bytes  # file1 - Takeoff
        main_file_name = file2_name
        takeoff_file_name = file1_name
    else:
        main_file_bytes = file1_bytes  # file1 - основной
        takeoff_file_bytes = file2_bytes  # file2 - Takeoff
        main_file_name = file1_name
        takeoff_file_name = file2_name
    
    print(f"Основной файл: {main_file_name}")
    print(f"Takeoff файл: {takeoff_file_name}")
    
    # Открываем PDF файлы
    doc1 = fitz.open(stream=main_file_bytes, filetype="pdf")
    doc2 = fitz.open(stream=takeoff_file_bytes, filetype="pdf")
    
    try:
        # === ЛИСТ 1: ОСНОВНОЕ ===
        print("Создаю лист 'Основное'...")
        lines = extract_first_n_lines_from_doc(doc1, n=32)
        while len(lines) < 32:
            lines.append("")

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Основное"

        ws1.cell(row=1, column=1, value=lines[0] if len(lines) > 0 else "")
        ws1.cell(row=2, column=1, value=lines[1] if len(lines) > 1 else "")
        ws1.cell(row=1, column=7, value=lines[2] if len(lines) > 2 else "")
        ws1.cell(row=2, column=7, value=lines[3] if len(lines) > 3 else "")

        block1 = lines[4:18]
        if len(block1) == 14:
            for col in range(7):
                ws1.cell(row=4, column=1 + col, value=block1[col * 2] if col * 2 < len(block1) else "")
                ws1.cell(row=5, column=1 + col, value=block1[col * 2 + 1] if col * 2 + 1 < len(block1) else "")

        block2 = lines[18:32]
        if len(block2) == 14:
            for col in range(7):
                ws1.cell(row=7, column=1 + col, value=block2[col * 2] if col * 2 < len(block2) else "")
                ws1.cell(row=8, column=1 + col, value=block2[col * 2 + 1] if col * 2 + 1 < len(block2) else "")

        bold_font = Font(bold=True)
        left_align = Alignment(horizontal="left", vertical="top")
        right_align = Alignment(horizontal="right", vertical="top")

        ws1['A1'].font = bold_font
        ws1['A1'].alignment = left_align
        ws1['G1'].alignment = right_align
        ws1['G2'].alignment = right_align

        for col in range(1, 8):
            ws1.cell(row=4, column=col).font = bold_font
            ws1.cell(row=4, column=col).alignment = left_align
            ws1.cell(row=7, column=col).font = bold_font
            ws1.cell(row=7, column=col).alignment = left_align

        for row in [2, 5, 8]:
            for col in range(1, 8):
                cell = ws1.cell(row=row, column=col)
                if cell.value is not None:
                    cell.alignment = left_align

        col_widths = [12, 11, 20, 14, 15, 10, 13]
        for i, w in enumerate(col_widths, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        ws1.page_setup.orientation = 'portrait'
        ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
        ws1.page_margins.left = 0.2
        ws1.page_margins.right = 0.2
        ws1.page_margins.top = 0.3
        ws1.page_margins.bottom = 0.3
        ws1.print_area = 'A1:G8'
        ws1.page_setup.fitToWidth = 1
        ws1.page_setup.fitToHeight = False

        # === ЛИСТ 2: ПАРСИНГ ТАБЛИЦЫ ===
        print("Парсинг таблицы маршрута...")
        df = parse_main_route_table(doc1)

        # Создание листа Main_Route_Grid
        ws2 = wb.create_sheet(title="Main_Route_Grid")

        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        align_center = Alignment(horizontal="center", vertical="center")

        # Заголовки
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws2.cell(row=2, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # Данные
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws2.cell(row=r_idx, column=c_idx, value=value)

        # Стилизация и объединение строки 1
        num_cols = len(df.columns)
        for col_idx in range(1, num_cols + 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # Объединение и значения
        ws2.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
        ws2.cell(row=1, column=3, value="MAG")

        ws2.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
        ws2.cell(row=1, column=6, value="WIND")

        ws2.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)
        ws2.cell(row=1, column=9, value="SPD KT")

        ws2.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12)
        ws2.cell(row=1, column=11, value="DIST NM")

        ws2.merge_cells(start_row=1, start_column=13, end_row=1, end_column=14)
        ws2.cell(row=1, column=13, value="FUEL G")

        ws2.merge_cells(start_row=1, start_column=16, end_row=1, end_column=18)
        ws2.cell(row=1, column=16, value="TIME")

        # Автоширина
        max_col = ws2.max_column
        for col_idx in range(1, max_col + 1):
            max_len = 0
            for row in ws2.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws2.max_row):
                cell = row[0]
                if hasattr(cell, 'value') and cell.value is not None:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_len + 2, 50)
            ws2.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # === ЛИСТ 3: AIRPORT TABLE ===
        print("Парсинг таблицы аэропортов...")
        df_airport = parse_airport_table(doc1)
        
        if df_airport.empty:
            print("⚠️ Таблица аэропортов не найдена, создаю пустой лист")
            ws3 = wb.create_sheet(title="Airport_Table")
            ws3['A1'] = "Таблица аэропортов не найдена в документе"
        else:
            ws3 = wb.create_sheet(title="Airport_Table")

            headers = ["", "AIRPORT", "ETA", "WX", "TWR/CTAF", "CLR", "GND", "ELEV", "RWY", "LONGEST"]
            if len(headers) > len(df_airport.columns):
                headers = headers[:len(df_airport.columns)]
            elif len(headers) < len(df_airport.columns):
                headers += [""] * (len(df_airport.columns) - len(headers))

            # Заголовки
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font_yellow = Font(bold=True)

            for col_num, value in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col_num, value=value)
                cell.font = bold_font_yellow
                cell.fill = yellow_fill

            # Данные
            for r_idx, row in enumerate(dataframe_to_rows(df_airport, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws3.cell(row=r_idx, column=c_idx, value=value)

            # Автоширина
            max_col = ws3.max_column
            for col_idx in range(1, max_col + 1):
                max_len = 0
                for row in ws3.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws3.max_row):
                    cell = row[0]
                    if hasattr(cell, 'value') and cell.value is not None:
                        try:
                            max_len = max(max_len, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = min(max_len + 2, 50)
                ws3.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # === ЛИСТ 4: AIRPORT MAPS ===
        print("Извлечение карт аэропортов...")
        text_A1, text_A28, img_buffers = extract_airport_maps(doc1)
        
        ws4 = wb.create_sheet(title="Airport_Maps")
        ws4.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.1, footer=0.1)

        ws4['A1'] = text_A1
        ws4['A1'].font = Font(bold=True)

        # Вставка изображений
        if len(img_buffers) >= 1:
            img_buffer = img_buffers[0]
            img_buffer.seek(0)
            ws4.add_image(XLImage(img_buffer), 'A2')
        
        if len(img_buffers) >= 2:
            img_buffer = img_buffers[1]
            img_buffer.seek(0)
            ws4.add_image(XLImage(img_buffer), 'A29')

        ws4['A28'] = text_A28
        ws4['A28'].font = Font(bold=True)

        ws4.column_dimensions['A'].width = 70

        # === ЛИСТ 5: ForeFlight (из Takeoff файла) ===
        print("Создание листа 'ForeFlight' из Takeoff файла...")
        
        page_ff = doc2[0]
        left_lines, right_lines = parse_document_with_simple_split(page_ff, "All Engines Operating")

        variables_1 = extract_variables(left_lines, "1")
        variables_2 = extract_variables(right_lines, "2")

        if "Runway1" in variables_1:
            runway01, new_runway1 = process_runway_variable(variables_1["Runway1"], "1")
            variables_1["Runway01"] = runway01
            variables_1["Runway1"] = new_runway1

        if "Runway2" in variables_2:
            runway02, new_runway2 = process_runway_variable(variables_2["Runway2"], "2")
            variables_2["Runway02"] = runway02
            variables_2["Runway2"] = new_runway2

        if "Wind1" in variables_1 and "Runway01" in variables_1:
            wind01, new_wind1 = process_wind_variable(variables_1["Wind1"], variables_1["Runway01"], "1")
            variables_1["Wind01"] = wind01
            variables_1["Wind1"] = new_wind1

        if "Wind2" in variables_2 and "Runway02" in variables_2:
            wind02, new_wind2 = process_wind_variable(variables_2["Wind2"], variables_2["Runway02"], "2")
            variables_2["Wind02"] = wind02
            variables_2["Wind2"] = new_wind2

        order_of_vars = [
            "Runway01", "Runway1", "Length1", "Surface1", "Wind01", "Wind1", "Altimeter1", "Distance1",
            "Runway02", "Runway2", "Length2", "Surface2", "Wind02", "Wind2", "Altimeter2", "Distance2"
        ]

        var_names_col3 = []
        var_values_col4 = []
        var_names_col5 = []
        var_values_col6 = []

        for var_name in order_of_vars:
            if any(var_name.endswith(suff) for suff in ["1", "01"]):
                if var_name in variables_1:
                    var_names_col3.append(var_name)
                    var_values_col4.append(variables_1[var_name])
                else:
                    var_names_col3.append("")
                    var_values_col4.append("")
            elif any(var_name.endswith(suff) for suff in ["2", "02"]):
                if var_name in variables_2:
                    var_names_col5.append(var_name)
                    var_values_col6.append(variables_2[var_name])
                else:
                    var_names_col5.append("")
                    var_values_col6.append("")

        max_var_len = max(len(var_names_col3), len(var_names_col5))

        var_names_col3 += [""] * (max_var_len - len(var_names_col3))
        var_values_col4 += [""] * (max_var_len - len(var_values_col4))
        var_names_col5 += [""] * (max_var_len - len(var_names_col5))
        var_values_col6 += [""] * (max_var_len - len(var_values_col6))

        df_vars = pd.DataFrame({
            'Variable_Name_1': var_names_col3,
            'Variable_Value_1': var_values_col4,
            'Variable_Name_2': var_names_col5,
            'Variable_Value_2': var_values_col6
        })

        max_len_arrays = max(len(left_lines), len(right_lines))
        left_extended = left_lines + [""] * (max_len_arrays - len(left_lines))
        right_extended = right_lines + [""] * (max_len_arrays - len(right_lines))

        df_arrays = pd.DataFrame({
            'Left_Column': left_extended,
            'Right_Column': right_extended
        })

        df_combined = pd.concat([df_arrays, df_vars], axis=1, sort=False).fillna("")

        ws5 = wb.create_sheet(title="ForeFlight")

        for r_idx, row in enumerate(dataframe_to_rows(df_combined, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws5.cell(row=r_idx, column=c_idx, value=value)

        for col in range(1, 7):
            ws5.column_dimensions[get_column_letter(col)].width = 25

        # === ЛИСТ 6: Generated_Sheet ===
        print("Создание листа 'Generated_Sheet'...")
        
        # Извлекаем скриншот для вставки
        screenshot_buffer = extract_screenshot_for_generated_sheet(doc1)
        
        # Создаем лист Generated_Sheet с вставкой скриншота
        create_generated_sheet(wb, ws1, ws2, ws3, ws4, ws5, screenshot_buffer)

        print(f"✅ Обработка завершена! Создан Excel файл с 6 листами.")
        
    finally:
        # Закрываем документы
        doc1.close()
        doc2.close()

    # Сохраняем workbook в bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes.getvalue()

def process(input_path, output_path):
    """Функция для обратной совместимости (не используется для двух файлов)"""
    raise NotImplementedError("Для обработки двух файлов используйте process_two_pdfs")

def main():
    """Для локального тестирования"""
    import sys
    if len(sys.argv) != 3:
        print("Использование: python your_script.py <файл1.pdf> <файл2.pdf>")
        sys.exit(1)
    
    with open(sys.argv[1], 'rb') as f1, open(sys.argv[2], 'rb') as f2:
        result = process_two_pdfs(f1.read(), f2.read(), sys.argv[1], sys.argv[2])
        
    with open("Flight_Log_Extracted.xlsx", 'wb') as f:
        f.write(result)
    print("Файл сохранен как Flight_Log_Extracted.xlsx")

if __name__ == "__main__":
    main()
